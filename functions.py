from openpyxl import load_workbook

EXCEL = 'Constructions.xlsx'
EXCEL1 = 'Construction ratios.xlsx'
wb = load_workbook(EXCEL)
wb1 = load_workbook(EXCEL1)
ws = wb.active
ws1 = wb1.active

WATCHER = 590000
CHAOS_DRAGON = 660000
letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
constructions = ['Academy',
                 'Altar',
                 'Barrack',
                 'Battle_Hall',
                 'Castle',
                 'Castle_Wall',
                 'Embassy',
                 'Farm',
                 'Gym',
                 'Infirmary',
                 'Lumber_Mill',
                 'Lunar_Foundry',
                 'Manor',
                 'Mine',
                 'Monsterhold',
                 'Mystic_Spire',
                 'Prison',
                 'Quarry',
                 'Spring',
                 'Trading_Post',
                 'Treasure_Trove',
                 'Vault',
                 'Watchtower',
                 'Workshop']
VIP_LEVELS = {
    1: 5,
    2: 7,
    3: 9,
    4: 11,
    5: 13,
    6: 15,
    7: 18,
    8: 21,
    9: 24,
    10: 27,
    11: 30,
    12: 35,
    13: 40,
    14: 45,
    15: 60
}

might_bonuses = []  # List of cells where "Might Bonus" is at each cell
original_times = []  # List of cells where "Original Time" is at each cell
food_list = []  # List of cells where "Food Cost" is at each cell
stone_list = []  # List of cells where "Stone Cost" is at each cell
timber_list = []  # List of cells where "Wood Cost" is at each cell
ore_list = []  # List of cells where "Ore Cost" is at each cell

might_table = []
time_table = []
food_table = []
stone_table = []
timber_table = []
ore_table = []

might_dict = {}
time_dict = {}
food_dict = {}
stone_dict = {}
timber_dict = {}
ore_dict = {}
rss_dict = {}

net_time = {}
helped_time = {}
VIP_time = {}
ratio_conversor = {}


def find_titles():  # Finds the cells whose title is interesting (Might Bonus, Original Time, Food Cost and so on)
    row = 1
    for i in range(0, wb.active.max_row + 1):
        sheet = str(wb.active).split("\"", 2)[1]
        column = 1
        for x in range(column, wb.active.max_column + 1):
            cell = wb.active[f'{translate_number(column)}{row}']
            if cell.value == 'Might Bonus':
                might_bonuses.append(cell.coordinate)
            if cell.value == 'Original Time':
                original_times.append(cell.coordinate)
            if cell.value == 'Food Cost':
                food_list.append(cell.coordinate)
            if cell.value == 'Stone Cost':
                stone_list.append(cell.coordinate)
            if cell.value == 'Timber Cost':
                timber_list.append(cell.coordinate)
            if cell.value == 'Ore Cost':
                ore_list.append(cell.coordinate)
            column += 1
        row += 1


def assign_mights(func_construction):  # It searches the 25 cells below the cell with the value "Might Bonus", which correspond to the original might earned when upgrading that building from level 1 to 25
    construction_index = int(constructions.index(func_construction))
    wb.active = construction_index
    mights = []
    might_cell = might_bonuses[construction_index]
    # finds the cell with "Might Bonus" in the sheet with the name dict_construction
    might_column = str(might_cell)[0]
    might_row = int(might_cell[1:])
    # separates the column and row to make iteration easier

    for i in range(0, 25):

        might_row += 1
        st_ring = wb.active[f'{might_column}{might_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        mights.append(clean_string)
    return mights


def assign_times(func_construction):  # It searches the 25 cells below the cell with the value "Original Time", which correspond to the original times for that building from level 1 to 25
    construction_index = int(constructions.index(func_construction))
    wb.active = construction_index
    times = []
    time_cell = original_times[construction_index]
    # finds the cell with "Original Time" in the sheet with the name original_times
    time_column = str(time_cell)[0]
    time_row = int(time_cell[1:])
    # separates the column and row to make iteration easier

    for i in range(0, 25):

        time_row += 1
        st_ring = wb.active[f'{time_column}{time_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        times.append(clean_string)
    return times


def assign_foods(func_construction):  # It searches the 25 cells below the cell with the value "Food Cost", which correspond to the food_list for that building from level 1 to 25
    foods = []
    construction_index = int(constructions.index(func_construction))
    wb.active = construction_index
    if int(constructions.index(func_construction)) == 7:  # Index 7 is "Farm", and it doesn't need food to improve, that messes with the indexes
        for i in range(0, 25):
            foods.append(None)
        return foods
    elif int(constructions.index(func_construction)) < 7:
        food_cell = food_list[construction_index]
    else:
        food_cell = food_list[construction_index - 1]
    food_column = str(food_cell)[0]
    food_row = int(food_cell[1:])
        # separates the column and row to make iteration easier
    for i in range(0, 25):
        food_row += 1
        st_ring = wb.active[f'{food_column}{food_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        foods.append(clean_string)
    return foods


def assign_stones(func_construction):  # It searches the 25 cells below the cell with the value "Food Cost", which correspond to the food_list for that building from level 1 to 25
    stones = []
    construction_index = int(constructions.index(func_construction))
    wb.active = construction_index
    if int(constructions.index(func_construction)) == 17:  # Index 17 is "Quarry", and it doesn't need stone to improve, that messes with the indexes
        for i in range(0, 25):
            stones.append(None)
        return stones
    elif int(constructions.index(func_construction)) < 17:
        stone_cell = stone_list[construction_index]
    else:
        stone_cell = stone_list[construction_index - 1]
    stone_column = str(stone_cell)[0]
    stone_row = int(stone_cell[1:])
        # separates the column and row to make iteration easier
    for i in range(0, 25):
        stone_row += 1
        st_ring = wb.active[f'{stone_column}{stone_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        stones.append(clean_string)
    return stones


def assign_timbers(dict_timber):  # It searches the 25 cells below the cell with the value "Food Cost", which correspond to the food_list for that building from level 1 to 25
    timbers = []
    construction_index = int(constructions.index(dict_timber))
    wb.active = construction_index
    if int(constructions.index(dict_timber)) == 10:  # Index 10 is "Lumber_Mill", and it doesn't need timber to improve, that messes with the indexes
        for i in range(0, 25):
            timbers.append(None)
        return timbers
    elif int(constructions.index(dict_timber)) < 10:
        timber_cell = timber_list[construction_index]
    else:
        timber_cell = timber_list[construction_index - 1]
    timber_column = str(timber_cell)[0]
    timber_row = int(timber_cell[1:])  # separates the column and row to make iteration easier
    for i in range(0, 25):
        timber_row += 1
        st_ring = wb.active[f'{timber_column}{timber_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        timbers.append(clean_string)
    return timbers


def assign_ores(dict_ore):  # It searches the 25 cells below the cell with the value "Food Cost", which correspond to the food_list for that building from level 1 to 25
    ores = []
    construction_index = int(constructions.index(dict_ore))
    wb.active = construction_index
    if int(constructions.index(dict_ore)) == 13:  # Index 13 is "Mine", and it doesn't need ore to improve, that messes with the indexes
        for i in range(0, 25):
            ores.append(None)
        return ores
    elif int(constructions.index(dict_ore)) < 13:
        ore_cell = ore_list[construction_index]
    else:
        ore_cell = ore_list[construction_index - 1]
    ore_column = str(ore_cell)[0]
    ore_row = int(ore_cell[1:])  # separates the column and row to make iteration easier
    for i in range(0, 25):
        ore_row += 1
        st_ring = wb.active[f'{ore_column}{ore_row}'].value
        try:
            clean_string = int(str(st_ring).replace(",", ""))
        except ValueError:
            clean_string = st_ring
        ores.append(clean_string)
    return ores


def translate_time(item):  # This translates strings like "1d 19:03:30" into seconds
    seconds = 0
    try:
        for i in item.split(" ", 1):
            if ':' not in i:
                seconds += int(i[:-1]) * 86400
            else:
                rounds = 0
                for split in i.split(":", 2):
                    seconds += int(split) * (60 ** (2 - rounds))
                    rounds += 1
        return seconds
    except ValueError:
        return item
    except AttributeError:
        return item


def translate_number(number):  # Translate an index into an excel column (only works from A to Z)
    letter = letters[number - 1]
    return letter


def add_gear(gear):  # Take into account the construction gear, which makes constructions take less time
    for i in time_dict:
        try:
            net_time[i] = round((int(time_dict[i]) / (float(gear) / 100 + 1)), 2)
        except (ValueError, TypeError):
            net_time[i] = time_dict[i]


def add_helps(helps=30):  # Take into account the guildmates' help, which reduce the time of the construction. It can vary, so I will assume you get the helps instantaneously (it doesn't vary so much from reality)
    for i in net_time:
        interim_time = net_time[i]
        for x in range(0, helps):
            try:
                if interim_time > max(round((net_time[i] * 0.01), 2), 60):
                    interim_time -= max(round((net_time[i] * 0.01), 2), 60)
                else:
                    interim_time = 0
            except (ValueError, TypeError):
                helped_time[i] = net_time[i]
        try:
            helped_time[i] = round(int(interim_time), 2)
        except (ValueError, TypeError):
            helped_time[i] = net_time[i]


def add_vip(VIP_LEVEL):  # Take into account the VIP level, which lets you finish constructions faster (when x minutes are left, it's free to speed up)
    seconds_saved = int(VIP_LEVELS[VIP_LEVEL] * 60)
    for i in helped_time:
        try:
            if int(helped_time[i]) < seconds_saved:
                VIP_time[i] = 0
            else:
                VIP_time[i] = round(int((helped_time[i]) - seconds_saved), 2)
        except (ValueError, TypeError):
            VIP_time[i] = helped_time[i]

