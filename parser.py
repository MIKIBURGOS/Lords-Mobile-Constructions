from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook


EXCEL = 'Constructions.xlsx'
wb = load_workbook(EXCEL)
ws = wb.active

constructions_index = 0
constructions = ['Academy',
                 'Altar',
                 'Barrack',
                 'Battle_Hall',
                 'Castle',
                 'Castle_Wall',
                 'Embassy',
                 'Farm',
                 'Gym',
                 'Infirmary',
                 'Lumber_Mill',
                 'Lunar_Foundry',
                 'Manor',
                 'Mine',
                 'Monsterhold',
                 'Mystic_Spire',
                 'Prison',
                 'Quarry',
                 'Spring',
                 'Trading_Post',
                 'Treasure_Trove',
                 'Vault',
                 'Watchtower',
                 'Workshop']

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']


def translate_number(number):
    letter = letters[number - 1]
    return letter

def create_sheets():
    for i in constructions:
        print(i)
        wb.create_sheet(f'{i}')
        print(wb.sheetnames)
        wb.save(EXCEL)


def parse():
    content = driver.page_source
    soup = BeautifulSoup(content, features="lxml")
    result_rows = []
    for table in soup.findAll('table', attrs={'class': 'article-table'}):
        rows = table.findAll('tr')
        # Recorrer las filas de la tabla
        for row in rows:
            # Array para meter los textos de las celdas de esta fila
            result_row = []
            # Recorrer las celdas de la fila
            for cell in row.select('td, th'):
                clean_text = cell.get_text().replace('\n', '')
                result_row.append(clean_text.replace('\xa0', ''))
            result_rows.append(result_row)

    row = 1
    for i in result_rows:
        column = 1
        for item in i:
            column_letter = translate_number(column)
            list_index = 0
            coords = f'{column_letter}{row}'
            wb.active[coords] = item
            list_index += 1
            column += 1
            wb.save(EXCEL)
        row += 1
        wb.save(EXCEL)


def send():
    send_index = 0
    for construction in constructions:
        wb.active = send_index
        print(wb.active)
        url = f'https://lordsmobile.fandom.com/wiki/{constructions[constructions_index]}'
        driver.get(url)
        parse()
        send_index += 1
        wb.save(EXCEL)





